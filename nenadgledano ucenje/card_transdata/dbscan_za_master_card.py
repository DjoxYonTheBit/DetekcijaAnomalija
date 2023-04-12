import pandas as pd
import numpy as np
from sklearn.cluster import DBSCAN
from sklearn import preprocessing, metrics
from openpyxl import load_workbook

df = pd.read_csv('card_transdata.csv')
df = df.sample(n=10000)
X = df.drop(['fraud'], axis=1)
y = df['fraud']

scaler = preprocessing.StandardScaler()
X = scaler.fit_transform(X)

eps = np.linspace(0.2, 4.0, num=8)
min_samples = range(3, 8)

best_score = -1
best_eps = -1
best_min_samples = -1

for e in eps:
    for m in min_samples:
        dbscan = DBSCAN(eps=e, min_samples=m).fit(X)
        score = metrics.silhouette_score(X, dbscan.labels_)
        if score > best_score:
            best_score = score
            best_eps = e
            best_min_samples = m

dbscan = DBSCAN(eps=best_eps, min_samples=best_min_samples).fit(X)
sil_skor=metrics.silhouette_score(X, dbscan.labels_)
print('Silhouette score:',sil_skor )

y_pred = np.where(dbscan.labels_ == -1, 1, 0)
tn, fp, fn, tp = metrics.confusion_matrix(y, y_pred).ravel()
print('Confusion matrix:')
print('TN =', tn, 'FP =', fp)
print('FN =', fn, 'TP =', tp)

# preciznost i odziv (rucno implementirani)
precision = tp / (tp + fp)
recall = tp / (tp + fn)
print('Precision:', precision)
print('Recall:', recall)

wb = load_workbook('izvjestaj_klaster_card.xlsx')

ws = wb.active

column_num = ws.max_column #+ 1

ws.cell(row=1, column=column_num, value='Metrika')
ws.cell(row=2, column=column_num, value='Precision')
ws.cell(row=3, column=column_num, value='Recall')
ws.cell(row=4, column=column_num, value='silhouette score')

ws.cell(row=1, column=column_num+1, value='DBSCAN')
ws.cell(row=2, column=column_num+1, value=precision)
ws.cell(row=3, column=column_num+1, value=recall)
ws.cell(row=4, column=column_num+1, value=sil_skor)

wb.save('izvjestaj_klaster_card.xlsx')
