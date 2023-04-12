import pandas as pd
from sklearn import preprocessing,metrics,ensemble
from openpyxl import load_workbook

df=pd.read_csv("card_transdata.csv")
df=df.sample(n=10000)
X=df.drop(['fraud'],axis=1)
y=df['fraud']

scaler=preprocessing.StandardScaler()
scaler.fit(X)
X=scaler.transform(X)

model=ensemble.IsolationForest(n_estimators=250,contamination=0.007)

model.fit(X)
y_pred=model.predict(X)
y_labele=y_pred

for i in range(0,len(y_pred)):
    if y_pred[i]==-1:
        y_pred[i]=1
    else: y_pred[i]=0

TP=sum((y_pred==1) & (y==1))
FP=sum((y_pred==1) & (y==0))
TN=sum((y_pred==0) & (y==0))
FN=sum((y_pred==0) & (y==1))

preciznost=TP/(TP+FP)
odziv=TP/(TP+FN)
sil_skor=metrics.silhouette_score(X,y_labele)
print("PRECIZNOST",preciznost)
print("ODZIV: ",odziv)
print("skor siluete",sil_skor)

#print("PRRRRRRRRRR",metrics.precision_score(y, y_pred))

wb = load_workbook('izvjestaj_klaster_card.xlsx')

ws = wb.active

column_num = ws.max_column #+ 1

ws.cell(row=1, column=column_num+1, value='Iforest')
ws.cell(row=2, column=column_num+1, value=preciznost)
ws.cell(row=3, column=column_num+1, value=odziv)
ws.cell(row=4, column=column_num+1, value=sil_skor)

wb.save('izvjestaj_klaster_card.xlsx')
