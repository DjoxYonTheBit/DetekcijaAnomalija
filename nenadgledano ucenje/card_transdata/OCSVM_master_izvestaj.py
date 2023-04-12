import pandas as pd
from sklearn import preprocessing, metrics
from sklearn.svm import OneClassSVM
from openpyxl import load_workbook

df = pd.read_csv("card_transdata.csv")
df = df.sample(n=10000)
X = df.drop(['fraud'], axis=1)
y = df['fraud']

scaler = preprocessing.StandardScaler()
scaler.fit(X)
X = scaler.transform(X)

model = OneClassSVM(kernel='rbf', nu=0.5)
model.fit(X)
y_pred = model.predict(X)
for i in range(0,len(y_pred)):
    if y_pred[i]==-1:
        y_pred[i]=1
    else: y_pred[i]=0

TP = sum((y_pred == 1) & (y == 1))
FP = sum((y_pred == 1) & (y == 0))
TN = sum((y_pred == 0) & (y == 0))
FN = sum((y_pred == 0) & (y == 1))

preciznost = TP / (TP + FP)
odziv = TP / (TP + FN)

print("PRECIZNOST:", preciznost)
print("ODZIV:", odziv)
print("Skor siluete:", metrics.silhouette_score(X, y_pred))
print("PRECIZNOST (2):", metrics.precision_score(y, y_pred))


wb = load_workbook('izvjestaj_klaster_card.xlsx')

ws = wb.active

column_num = ws.max_column #+ 1

ws.cell(row=1, column=column_num+1, value='OC-SVM')
ws.cell(row=2, column=column_num+1, value=preciznost)
ws.cell(row=3, column=column_num+1, value=odziv)

wb.save('izvjestaj_klaster_card.xlsx')
